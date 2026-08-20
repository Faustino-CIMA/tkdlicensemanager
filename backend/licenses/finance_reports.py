from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO

from django.db.models import Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from clubs.models import FederationProfile

from .models import Expense, FinanceYearOpening, Invoice, Payment

ZERO = Decimal("0.00")
MONEY_QUANT = Decimal("0.01")


def money(value) -> Decimal:
    if value is None:
        return ZERO
    return Decimal(value).quantize(MONEY_QUANT)


def money_str(value) -> str:
    return f"{money(value):.2f}"


def year_as_of(year: int, today: date | None = None) -> date:
    resolved_today = today or timezone.localdate()
    year_end = date(year, 12, 31)
    if year == resolved_today.year:
        return resolved_today
    if year > resolved_today.year:
        return resolved_today
    return year_end


def period_bounds(year: int, as_of: date):
    start = date(year, 1, 1)
    if as_of < start:
        as_of = start
    start_dt = timezone.make_aware(datetime.combine(start, time.min))
    as_of_end_dt = timezone.make_aware(datetime.combine(as_of, time.max))
    return start, as_of, start_dt, as_of_end_dt


def _invoice_recognition_q(start_dt, as_of_end_dt):
    issued_in_period = Q(issued_at__gte=start_dt, issued_at__lte=as_of_end_dt)
    created_fallback = Q(issued_at__isnull=True, created_at__gte=start_dt, created_at__lte=as_of_end_dt)
    return issued_in_period | created_fallback


def _invoice_recognized_by(as_of_end_dt):
    issued_by = Q(issued_at__lte=as_of_end_dt)
    created_fallback = Q(issued_at__isnull=True, created_at__lte=as_of_end_dt)
    return issued_by | created_fallback


def _paid_at_in_period(start_dt, as_of_end_dt):
    paid_in_period = Q(paid_at__gte=start_dt, paid_at__lte=as_of_end_dt)
    created_fallback = Q(paid_at__isnull=True, created_at__gte=start_dt, created_at__lte=as_of_end_dt)
    return paid_in_period | created_fallback


def _paid_at_on_or_before(as_of_end_dt):
    return Q(paid_at__lte=as_of_end_dt) | Q(paid_at__isnull=True, created_at__lte=as_of_end_dt)


def _paid_at_before(start_dt):
    return Q(paid_at__lt=start_dt) | Q(paid_at__isnull=True, created_at__lt=start_dt)


def _sum(queryset, field="amount"):
    return money(queryset.aggregate(total=Sum(field))["total"])


def computed_opening_cash(start_dt) -> Decimal:
    receipts_before = Payment.objects.filter(status=Payment.Status.PAID).filter(_paid_at_before(start_dt))
    expenses_before = Expense.objects.filter(status=Expense.Status.PAID).filter(_paid_at_before(start_dt))
    return _sum(receipts_before) - _sum(expenses_before)


def resolve_opening_cash(year: int, start_dt):
    stored = FinanceYearOpening.objects.filter(year=year).first()
    if stored:
        return money(stored.opening_cash), True, stored.notes
    return computed_opening_cash(start_dt), False, ""


def build_finance_report(year: int, today: date | None = None) -> dict:
    as_of = year_as_of(year, today)
    start, as_of, start_dt, as_of_end_dt = period_bounds(year, as_of)
    profile = FederationProfile.objects.order_by("id").first()
    organization_name = profile.name if profile else "Luxembourg Taekwondo Federation"

    revenue_invoices = Invoice.objects.filter(
        status__in=[Invoice.Status.ISSUED, Invoice.Status.PAID]
    ).filter(_invoice_recognition_q(start_dt, as_of_end_dt))
    revenue_total = _sum(revenue_invoices, "total")
    income_by_club = [
        {
            "club_id": row["club_id"],
            "club_name": row["club__name"] or f"Club {row['club_id']}",
            "amount": money_str(row["total"]),
        }
        for row in revenue_invoices.values("club_id", "club__name")
        .annotate(total=Sum("total"))
        .order_by("club__name")
    ]

    period_expenses = Expense.objects.filter(
        status__in=[Expense.Status.RECORDED, Expense.Status.PAID],
        expense_date__gte=start,
        expense_date__lte=as_of,
    ).select_related("category", "club")
    expense_total = _sum(period_expenses)
    expenses_by_category = [
        {
            "category_id": row["category_id"],
            "category_code": row["category__code"],
            "category_name": row["category__name"],
            "amount": money_str(row["total"]),
        }
        for row in period_expenses.values("category_id", "category__code", "category__name")
        .annotate(total=Sum("amount"))
        .order_by("category__sort_order", "category__name")
    ]
    expense_register = [
        {
            "id": expense.id,
            "expense_number": expense.expense_number,
            "expense_date": expense.expense_date.isoformat(),
            "category_name": expense.category.name,
            "payee": expense.payee,
            "description": expense.description,
            "amount": money_str(expense.amount),
            "status": expense.status,
            "club_name": expense.club.name if expense.club_id else "",
            "paid_at": expense.paid_at.isoformat().replace("+00:00", "Z") if expense.paid_at else None,
            "reference": expense.reference,
        }
        for expense in period_expenses.order_by("expense_date", "id")
    ]

    surplus = revenue_total - expense_total

    opening_cash, opening_is_manual, opening_notes = resolve_opening_cash(year, start_dt)
    receipts = _sum(
        Payment.objects.filter(status=Payment.Status.PAID).filter(_paid_at_in_period(start_dt, as_of_end_dt))
    )
    disbursements = _sum(
        Expense.objects.filter(status=Expense.Status.PAID).filter(_paid_at_in_period(start_dt, as_of_end_dt))
    )
    closing_cash = opening_cash + receipts - disbursements

    receivables_qs = (
        Invoice.objects.filter(_invoice_recognized_by(as_of_end_dt))
        .filter(status__in=[Invoice.Status.ISSUED, Invoice.Status.PAID])
        .exclude(Q(status=Invoice.Status.PAID) & _paid_at_on_or_before(as_of_end_dt))
    )
    receivables = [
        {
            "id": invoice.id,
            "invoice_number": invoice.invoice_number,
            "club_name": invoice.club.name if invoice.club_id else "",
            "issued_at": (invoice.issued_at or invoice.created_at).isoformat().replace("+00:00", "Z"),
            "amount": money_str(invoice.total),
        }
        for invoice in receivables_qs.select_related("club").order_by("issued_at", "id")
    ]
    accounts_receivable = _sum(receivables_qs, "total")

    payables_qs = Expense.objects.filter(
        expense_date__lte=as_of,
        status__in=[Expense.Status.RECORDED, Expense.Status.PAID],
    ).exclude(Q(status=Expense.Status.PAID) & _paid_at_on_or_before(as_of_end_dt))
    payables = [
        {
            "id": expense.id,
            "expense_number": expense.expense_number,
            "payee": expense.payee,
            "description": expense.description,
            "expense_date": expense.expense_date.isoformat(),
            "amount": money_str(expense.amount),
        }
        for expense in payables_qs.select_related("category").order_by("expense_date", "id")
    ]
    accounts_payable = _sum(payables_qs)

    total_assets = closing_cash + accounts_receivable
    total_liabilities = accounts_payable
    net_assets = total_assets - total_liabilities

    return {
        "organization_name": organization_name,
        "currency": "EUR",
        "year": year,
        "period_start": start.isoformat(),
        "as_of": as_of.isoformat(),
        "generated_at": timezone.now().isoformat().replace("+00:00", "Z"),
        "methodology": (
            "Accrual books for the selected calendar year. License income is recognized when "
            "an invoice is issued. Expenses are recognized on their expense date. The balance "
            "sheet cash figure is opening cash plus receipts minus paid expenses through the "
            "as-of date. Receivables are issued invoices still unpaid at that date. Payables "
            "are recorded expenses still unpaid at that date."
        ),
        "opening": {
            "cash": money_str(opening_cash),
            "is_manual": opening_is_manual,
            "notes": opening_notes,
        },
        "income_statement": {
            "revenue_license_fees": money_str(revenue_total),
            "expenses_total": money_str(expense_total),
            "surplus": money_str(surplus),
            "income_by_club": income_by_club,
            "expenses_by_category": expenses_by_category,
        },
        "cash_movement": {
            "opening_cash": money_str(opening_cash),
            "receipts": money_str(receipts),
            "disbursements": money_str(disbursements),
            "closing_cash": money_str(closing_cash),
        },
        "balance_sheet": {
            "assets": {
                "cash": money_str(closing_cash),
                "accounts_receivable": money_str(accounts_receivable),
                "total": money_str(total_assets),
            },
            "liabilities": {
                "accounts_payable": money_str(accounts_payable),
                "total": money_str(total_liabilities),
            },
            "equity": {
                "net_assets": money_str(net_assets),
                "total": money_str(net_assets),
            },
            "liabilities_and_equity_total": money_str(total_liabilities + net_assets),
        },
        "registers": {
            "expenses": expense_register,
            "receivables": receivables,
            "payables": payables,
        },
    }


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=16, color="1F4E79")
SECTION_FONT = Font(bold=True, name="Calibri", size=12, color="1F4E79")
LABEL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(bold=True, name="Calibri", size=11)
MONEY_FORMAT = '#,##0.00'
THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
TOTAL_FILL = PatternFill("solid", fgColor="E8EEF4")


def _write_header_row(ws, row, headers, start_col=1):
    for index, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN


def _money_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=float(money(value)))
    cell.number_format = MONEY_FORMAT
    cell.font = LABEL_FONT
    cell.border = THIN
    cell.alignment = Alignment(horizontal="right")
    return cell


def _text_cell(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BOLD_FONT if bold else LABEL_FONT
    cell.border = THIN
    return cell


def _autosize(ws, min_width=12, max_width=42):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def render_finance_report_xlsx(report: dict) -> bytes:
    wb = Workbook()

    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = report["organization_name"]
    cover["A1"].font = TITLE_FONT
    cover["A2"] = f"Annual financial report {report['year']}"
    cover["A2"].font = SECTION_FONT
    cover["A4"] = "Period start"
    cover["B4"] = report["period_start"]
    cover["A5"] = "As of"
    cover["B5"] = report["as_of"]
    cover["A6"] = "Currency"
    cover["B6"] = report["currency"]
    cover["A7"] = "Generated"
    cover["B7"] = report["generated_at"]
    cover["A9"] = "How these figures are prepared"
    cover["A9"].font = SECTION_FONT
    cover["A10"] = report["methodology"]
    cover["A10"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.merge_cells("A10:F13")
    cover.row_dimensions[10].height = 72
    cover["A15"] = "Surplus / (deficit)"
    _money_cell(cover, 15, 2, report["income_statement"]["surplus"])
    cover["A16"] = "Cash at period end"
    _money_cell(cover, 16, 2, report["balance_sheet"]["assets"]["cash"])
    cover["A17"] = "Net assets"
    _money_cell(cover, 17, 2, report["balance_sheet"]["equity"]["net_assets"])
    _autosize(cover)

    pnl = wb.create_sheet("Income statement")
    pnl["A1"] = "Income statement"
    pnl["A1"].font = TITLE_FONT
    pnl["A2"] = f"{report['organization_name']} · {report['period_start']} to {report['as_of']}"
    _write_header_row(pnl, 4, ["Line", "Amount (EUR)"])
    _text_cell(pnl, 5, 1, "License fee income")
    _money_cell(pnl, 5, 2, report["income_statement"]["revenue_license_fees"])
    _text_cell(pnl, 6, 1, "Operating expenses")
    _money_cell(pnl, 6, 2, report["income_statement"]["expenses_total"])
    _text_cell(pnl, 7, 1, "Surplus / (deficit)", bold=True)
    total_cell = _money_cell(pnl, 7, 2, report["income_statement"]["surplus"])
    total_cell.font = BOLD_FONT
    pnl["A5"].fill = PatternFill("solid", fgColor="F6F8FA")
    pnl["A7"].fill = TOTAL_FILL
    pnl["B7"].fill = TOTAL_FILL
    pnl["A9"] = "Income by club"
    pnl["A9"].font = SECTION_FONT
    _write_header_row(pnl, 10, ["Club", "Amount (EUR)"])
    row = 11
    for item in report["income_statement"]["income_by_club"]:
        _text_cell(pnl, row, 1, item["club_name"])
        _money_cell(pnl, row, 2, item["amount"])
        row += 1
    if not report["income_statement"]["income_by_club"]:
        _text_cell(pnl, row, 1, "No license income in this period")
        row += 1
    row += 1
    pnl.cell(row=row, column=1, value="Expenses by category").font = SECTION_FONT
    row += 1
    _write_header_row(pnl, row, ["Category", "Amount (EUR)"])
    row += 1
    for item in report["income_statement"]["expenses_by_category"]:
        _text_cell(pnl, row, 1, item["category_name"])
        _money_cell(pnl, row, 2, item["amount"])
        row += 1
    if not report["income_statement"]["expenses_by_category"]:
        _text_cell(pnl, row, 1, "No expenses in this period")
    _autosize(pnl)

    bs = wb.create_sheet("Balance sheet")
    bs["A1"] = "Balance sheet"
    bs["A1"].font = TITLE_FONT
    bs["A2"] = f"{report['organization_name']} as of {report['as_of']}"
    _write_header_row(bs, 4, ["Assets", "Amount (EUR)"])
    _text_cell(bs, 5, 1, "Cash and bank")
    _money_cell(bs, 5, 2, report["balance_sheet"]["assets"]["cash"])
    _text_cell(bs, 6, 1, "Accounts receivable (unpaid invoices)")
    _money_cell(bs, 6, 2, report["balance_sheet"]["assets"]["accounts_receivable"])
    _text_cell(bs, 7, 1, "Total assets", bold=True)
    _money_cell(bs, 7, 2, report["balance_sheet"]["assets"]["total"]).font = BOLD_FONT
    bs["A7"].fill = TOTAL_FILL
    bs["B7"].fill = TOTAL_FILL
    _write_header_row(bs, 9, ["Liabilities", "Amount (EUR)"])
    _text_cell(bs, 10, 1, "Accounts payable (unpaid expenses)")
    _money_cell(bs, 10, 2, report["balance_sheet"]["liabilities"]["accounts_payable"])
    _text_cell(bs, 11, 1, "Total liabilities", bold=True)
    _money_cell(bs, 11, 2, report["balance_sheet"]["liabilities"]["total"]).font = BOLD_FONT
    _write_header_row(bs, 13, ["Equity / net assets", "Amount (EUR)"])
    _text_cell(bs, 14, 1, "Net assets")
    _money_cell(bs, 14, 2, report["balance_sheet"]["equity"]["net_assets"])
    _text_cell(bs, 16, 1, "Total liabilities and equity", bold=True)
    _money_cell(bs, 16, 2, report["balance_sheet"]["liabilities_and_equity_total"]).font = BOLD_FONT
    bs["A16"].fill = TOTAL_FILL
    bs["B16"].fill = TOTAL_FILL
    bs["A18"] = "Cash movement"
    bs["A18"].font = SECTION_FONT
    _write_header_row(bs, 19, ["Line", "Amount (EUR)"])
    _text_cell(bs, 20, 1, "Opening cash")
    _money_cell(bs, 20, 2, report["cash_movement"]["opening_cash"])
    _text_cell(bs, 21, 1, "Receipts (payments received)")
    _money_cell(bs, 21, 2, report["cash_movement"]["receipts"])
    _text_cell(bs, 22, 1, "Disbursements (expenses paid)")
    _money_cell(bs, 22, 2, report["cash_movement"]["disbursements"])
    _text_cell(bs, 23, 1, "Closing cash", bold=True)
    _money_cell(bs, 23, 2, report["cash_movement"]["closing_cash"]).font = BOLD_FONT
    _autosize(bs)

    expenses_sheet = wb.create_sheet("Expense register")
    expenses_sheet["A1"] = "Expense register"
    expenses_sheet["A1"].font = TITLE_FONT
    _write_header_row(
        expenses_sheet,
        3,
        ["Date", "Number", "Category", "Payee", "Description", "Amount (EUR)", "Status", "Club", "Reference"],
    )
    row = 4
    for item in report["registers"]["expenses"]:
        _text_cell(expenses_sheet, row, 1, item["expense_date"])
        _text_cell(expenses_sheet, row, 2, item["expense_number"])
        _text_cell(expenses_sheet, row, 3, item["category_name"])
        _text_cell(expenses_sheet, row, 4, item["payee"])
        _text_cell(expenses_sheet, row, 5, item["description"])
        _money_cell(expenses_sheet, row, 6, item["amount"])
        _text_cell(expenses_sheet, row, 7, item["status"])
        _text_cell(expenses_sheet, row, 8, item["club_name"])
        _text_cell(expenses_sheet, row, 9, item["reference"])
        row += 1
    if not report["registers"]["expenses"]:
        _text_cell(expenses_sheet, row, 1, "No expenses in this period")
    expenses_sheet.auto_filter.ref = f"A3:I{max(row - 1, 3)}"
    expenses_sheet.freeze_panes = "A4"
    _autosize(expenses_sheet)

    ar_sheet = wb.create_sheet("Receivables")
    ar_sheet["A1"] = "Open receivables"
    ar_sheet["A1"].font = TITLE_FONT
    ar_sheet["A2"] = f"Issued invoices unpaid as of {report['as_of']}"
    _write_header_row(ar_sheet, 4, ["Invoice", "Club", "Issued", "Amount (EUR)"])
    row = 5
    for item in report["registers"]["receivables"]:
        _text_cell(ar_sheet, row, 1, item["invoice_number"])
        _text_cell(ar_sheet, row, 2, item["club_name"])
        _text_cell(ar_sheet, row, 3, item["issued_at"][:10] if item["issued_at"] else "")
        _money_cell(ar_sheet, row, 4, item["amount"])
        row += 1
    if not report["registers"]["receivables"]:
        _text_cell(ar_sheet, row, 1, "No open receivables")
    _autosize(ar_sheet)

    ap_sheet = wb.create_sheet("Payables")
    ap_sheet["A1"] = "Open payables"
    ap_sheet["A1"].font = TITLE_FONT
    ap_sheet["A2"] = f"Recorded expenses unpaid as of {report['as_of']}"
    _write_header_row(ap_sheet, 4, ["Number", "Date", "Payee", "Description", "Amount (EUR)"])
    row = 5
    for item in report["registers"]["payables"]:
        _text_cell(ap_sheet, row, 1, item["expense_number"])
        _text_cell(ap_sheet, row, 2, item["expense_date"])
        _text_cell(ap_sheet, row, 3, item["payee"])
        _text_cell(ap_sheet, row, 4, item["description"])
        _money_cell(ap_sheet, row, 5, item["amount"])
        row += 1
    if not report["registers"]["payables"]:
        _text_cell(ap_sheet, row, 1, "No open payables")
    _autosize(ap_sheet)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
